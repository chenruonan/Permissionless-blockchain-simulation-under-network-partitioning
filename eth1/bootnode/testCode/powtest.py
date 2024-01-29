from web3 import Web3, EthereumTesterProvider
from web3.middleware import geth_poa_middleware
import web3
import solcx
import json
import time
import re
from solcx import install_solc
import random
import threading
import time
from web3.contract import Contract
from datetime import datetime,timedelta
import openpyxl
import subprocess
import os

# 解锁账户使之能部署合约
account_password = "1"  # 设置新账户的密码
# 解锁交换机1下的账户
def unlock_account_task(node_list:list):
    for node in node_list:
        unlock=node.geth.personal.unlock_account(node.eth.accounts[0],account_password,0)
        print(unlock,end=' ')
    print(f'{len(node_list)} node account unlock completed')

# deploy contract: input the node for deploying contract
def deployContract(node: Web3):
    with open("/usr/local/myproj/ethpowLab/testCode/erc20_ds.solidi") as f:
        contract_src_code = f.read()
        compiled_sol = solcx.compile_source(
        contract_src_code, output_values=["abi", "bin"], solc_version="0.8.19"
    )
    contract_id, contract_interface = compiled_sol.popitem()
    bytecode = contract_interface["bin"]
    abi = contract_interface["abi"]
    node.eth.default_account = node.eth.accounts[0]
    MyContract = node.eth.contract(abi=abi, bytecode=bytecode)
    tx_hash = MyContract.constructor().transact()
    tx_receipt = node.eth.wait_for_transaction_receipt(tx_hash)
    print("Contract deployed at address: ", tx_receipt.contractAddress)
    return tx_receipt.contractAddress, abi



if __name__=='__main__':
     # create Web3 object connecting to the node(Geth) of  execution layer
    switch1_node_addr=[] # 交换机1下节点的地址列表
    switch1_node_list=[] # 交换机1下节点的Web3对象列表
    for i in range(55):
        switch1_node_addr.append("http://192.168.1."+str(i+1)+":"+str(6000))
        switch1_node_list.append(Web3(web3.HTTPProvider(switch1_node_addr[i])))
    switch2_node_addr=[]
    switch2_node_list=[]
    for i in range(8):
        switch2_node_addr.append("http://192.168.2."+str(i+1)+":"+str(6000))
        switch2_node_list.append(Web3(web3.HTTPProvider(switch2_node_addr[i])))
    switch3_node_addr=[]
    switch3_node_list=[]
    for i in range(1):
        switch3_node_addr.append("http://192.168.3."+str(i+1)+":"+str(6000))
        switch3_node_list.append(Web3(web3.HTTPProvider(switch3_node_addr[i])))
    total_node_list=[]
    total_node_list.extend(switch1_node_list)
    total_node_list.extend(switch2_node_list)
    total_node_list.extend(switch3_node_list)
    for i in total_node_list:
        print(i.is_connected(),end=' ')
    print()

    unlock_thread1=threading.Thread(target=unlock_account_task,args=(total_node_list[0:12],))
    unlock_thread1.start()
    unlock_thread2=threading.Thread(target=unlock_account_task,args=(total_node_list[12:24],))
    unlock_thread2.start()
    unlock_thread3=threading.Thread(target=unlock_account_task,args=(total_node_list[24:36],))
    unlock_thread3.start()
    unlock_thread4=threading.Thread(target=unlock_account_task,args=(total_node_list[36:48],))
    unlock_thread4.start()
    unlock_thread5=threading.Thread(target=unlock_account_task,args=(total_node_list[48:64],))
    unlock_thread5.start()
    unlock_thread1.join()
    unlock_thread2.join()
    unlock_thread3.join()
    unlock_thread4.join()
    unlock_thread5.join()
    # deploy contract and send some money to the four nodes in the network
    total_contract_list=[]
    contract_address, abi = deployContract(total_node_list[0])
    time.sleep(3)
    for node in total_node_list:
        node.eth.default_account=node.eth.accounts[0]
        contract=node.eth.contract(address=contract_address,abi=abi)
        total_contract_list.append(contract)

    result = subprocess.run(["/usr/local/myproj/ethpowLab/testCode/adjustBandwidth.sh"], capture_output=True, text=True)
    # 输出脚本的标准输出
    print(result.stdout)

    # 输出脚本的标准错误
    print(result.stderr)
    time.sleep(2)

    workbook = openpyxl.load_workbook("/usr/local/myproj/ethpowLab/testCode/dsTest55_8.xlsx")
    sheet_name="512kbit"
            # 检查工作表是否已经存在
    if sheet_name in workbook.sheetnames:
            # 如果存在，直接使用该工作表
        ws = workbook[sheet_name]
    else:
            # 如果不存在，创建一个新的工作表
        ws = workbook.create_sheet(sheet_name, 0)

    startTime=datetime.now()
    for i in range(100):
        total_contract_list[0].functions.decrementByThree().transact()# 主网操作
        total_contract_list[57].functions.decrementByFour().transact()# 子网操作
    while(1):
        counterInMain=total_contract_list[0].functions.getCounter().call()
        counterInSub=total_contract_list[57].functions.getCounter().call()
        if datetime.now()-startTime>timedelta(minutes=3) and (counterInMain!=9300 or counterInSub!=9300) :
            ws.append(
                ( counterInMain,
                counterInSub,
                1, #代表发生双花了
                (datetime.now()-startTime).seconds,
                )
            )
            workbook.save("/usr/local/myproj/ethpowLab/testCode/dsTest55_8.xlsx")
            workbook.close()
            break
        if counterInMain==9300 and counterInSub==9300:
            print(f'主网计数器:{total_contract_list[0].functions.getCounter().call()}  子网计数器:{total_contract_list[57].functions.getCounter().call()}')
            ws.append(
                (counterInMain,
                counterInSub,
                0, #代表没有双花
                (datetime.now()-startTime).seconds,
                )
            )
            workbook.save("/usr/local/myproj/ethpowLab/testCode/dsTest55_8.xlsx")
            workbook.close()
            break
        else:
            print(f'主网计数器:{total_contract_list[0].functions.getCounter().call()}  子网计数器:{total_contract_list[57].functions.getCounter().call()}')
            time.sleep(2)

